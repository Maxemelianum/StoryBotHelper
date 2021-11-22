from telegram import Update
from telegram.ext import Updater, MessageHandler, Filters, CommandHandler, CallbackContext
from openpyxl import load_workbook

TOKEN = '2132309010:AAFQMIsoBEGKZYIR5kKy__dCI9keXwCIWKw'
book = load_workbook("-COPY - Copy.xlsx")

sheet_1 = book['Лист1']
sheet_2 = book['Лист2']
sheet_3 = book['Лист3']
sheet_4 = book['Лист4']
sheet_5 = book['Лист5']
sheet_6 = book['Лист6']
sheet_7 = book['Лист7']
sheet_8 = book['Лист8']
sheet_9 = book['Лист9']
sheet_10 = book['Лист10']
sheet_11 = book['Лист11']




def main():
    updater = Updater(token=TOKEN)  # объект, который ловит сообщения от Telegrama

    dispather = updater.dispatcher

    echo = MessageHandler(Filters.all, do_setcord)

    dispather.add_handler(echo)

    updater.start_polling()
    updater.idle()


def do_setcord(update, context: CallbackContext):
    if update.message.text == '/addvalue':
        context.user_data['command'] = update.message.text
        update.message.reply_text('Введите координаты ячейки:')

        if 'command' in context.user_data:
            if update.message.text != '/addvalue' and context.user_data['command'] == '/addvalue':
                do_setvalue(update, context)

        else:
            update.message.reply_text('chego?')


def do_setvalue(update: Update, context: CallbackContext):
    if context.user_data['command'] == '/addvalue':
        do_addvalue(update, context)

    else:
        update.message.reply_text(text='sry...')

        context.user_data['command'] = update.message.text
        return context.user_data['command']


def do_addvalue(update: Update, context):

    for i in range(2, 100):

        if sheet_7.cell(column=2, row=i).value is None:
            sheet_7.cell(column=2, row=i).value = update.message.text
            update.message.reply_text('Готово!')
            break

    return book.save('-COPY - Copy.xlsx')


main()