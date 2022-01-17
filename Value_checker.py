from openpyxl import load_workbook


book = load_workbook("DATAofStoryBotHelper.xlsx")

for page in book.sheetnames:
    print(page)

print('Введите название листа')

page = book[input()]

COLUMN_THEME = 1
COLUMN_MATERIAL = 3
COLUMN_MATERIAL_LINK = 4
COLUMN_TEST = 5
COLUMN_TEST_LINK = 6

reply1 = []

for line in range(2, page.max_row + 1):
    reply1.append(page.cell(row=line, column=COLUMN_THEME).value)

reply1 = set(reply1)
print(reply1)
answer1 = input('Введите тему:')
answer2 = input('Материалы или тесты?')
reply2 = []

for line in range(2, page.max_row + 1):
    value = page.cell(row=line, column=COLUMN_THEME).value
    if value == answer1:
        reply2.append(page.cell(row=line, column=COLUMN_MATERIAL).value)
print(reply2)

